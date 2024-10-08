from openpyxl import load_workbook
from datetime import datetime, timedelta
from tkinter import messagebox
import config


# Read holiday calendar to check people unavailability or holidays (find keywords in table)
def check_people_fields(file_path, keyword_list):
    wb = load_workbook(file_path)
    try:
        sheet = wb[config.HOLIDAY_SHEET_NAME]
    except KeyError:
        messagebox.showinfo("Error", f"Worksheet '{config.HOLIDAY_SHEET_NAME}' in '{config.HOLIDAY_FILE_NAME}' does "
                                     f"not exist")
        return

    holiday_dict = {}
    now = datetime.now()
    lower_bound = now - timedelta(days=config.HOLIDAY_MONTHS_SEARCH_RANGE * 30)
    upper_bound = now + timedelta(days=config.HOLIDAY_MONTHS_SEARCH_RANGE * 30)

    # Searching day by day
    for row in range(config.HOLIDAY_START_ROW, config.HOLIDAY_END_ROW + 1):
        date_cell = sheet.cell(row=row, column=config.HOLIDAY_START_COL)
        date_value = date_cell.value

        if date_value is None:
            continue

        # Date conversion "day/month/year"
        date_str = date_value.strftime('%d/%m/%Y')

        # Look for unavailability only in selected time range
        if lower_bound <= date_value <= upper_bound:
            # Searching in each column for selected row
            for col in range(config.HOLIDAY_NAMES_START_COL, config.HOLIDAY_END_COL + 1):
                person_name = sheet.cell(row=4, column=col).value
                availability = sheet.cell(row=row, column=col).value

                for keyword in keyword_list:
                    if availability == keyword:
                        # Add person to list
                        if date_str not in holiday_dict:
                            holiday_dict[date_str] = []
                        holiday_dict[date_str].append(person_name)

    wb.close()

    return holiday_dict
