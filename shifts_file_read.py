from openpyxl import load_workbook
import datetime


def count_weekend_days(file_path, shift_names):
    wb = load_workbook(file_path)
    # Search for sheet names with current year in title
    sheets_to_search = [sheet for sheet in wb.sheetnames if datetime.datetime.now().strftime('%Y') in sheet]
    # Create dictionary {name: [number of WSat (working saturdays), number of WSun (working sundays)]}
    results_dict = {name: [0, 0] for name in shift_names}

    # Searching in sheets which have selected (in config) year in name
    for sheet_name in sheets_to_search:
        sheet = wb[sheet_name]
        # Check column A if name matches shift_name in names_initials.csv
        for row in range(1, sheet.max_row + 1):
            person_name = sheet.cell(row=row, column=1).value
            if person_name in shift_names:
                # Look for working weekend days (Wsat or Wsun) from col B to the last to count them
                for col in range(2, sheet.max_column + 1):
                    shift_value = sheet.cell(row=row, column=col).value
                    if str(shift_value).lower() == 'wsat':
                        results_dict[person_name][0] += 1
                    elif str(shift_value).lower() == 'wsun':
                        results_dict[person_name][1] += 1
    wb.close()

    return results_dict
