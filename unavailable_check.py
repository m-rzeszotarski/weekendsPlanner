from openpyxl import load_workbook
from datetime import datetime, timedelta
import config


def check_people_availability(file_path):
    wb = load_workbook(file_path)
    sheet = wb['USP scheduling']

    unavailable_dict = {}
    now = datetime.now()
    lower_bound = now - timedelta(days=config.HOLIDAY_MONTHS_SEARCH_RANGE * 30)
    upper_bound = now + timedelta(days=config.HOLIDAY_MONTHS_SEARCH_RANGE * 30)

    # Searching day by day
    for row in range(config.HOLIDAY_START_ROW, config.HOLIDAY_END_ROW + 1):
        date_cell = sheet.cell(row=row, column=config.HOLIDAY_START_COL)
        date_value = date_cell.value

        if date_value is None:
            continue

        # Date conversion "month/day/year"
        date_str = date_value.strftime('%d/%m/%Y')

        # Look for unavailability only in selected time range
        if lower_bound <= date_value <= upper_bound:
            # Searching in each column for selected row
            for col in range(config.HOLIDAY_NAMES_START_COL, config.HOLIDAY_END_COL + 1):
                person_name = sheet.cell(row=4, column=col).value
                availability = sheet.cell(row=row, column=col).value

                for unavailable_word in config.UNAVAILABLE_WORDS_LIST:
                    if availability == unavailable_word:
                        # Add person to list
                        if date_str not in unavailable_dict:
                            unavailable_dict[date_str] = []
                        unavailable_dict[date_str].append(person_name)

    wb.close()

    return unavailable_dict
