from openpyxl import load_workbook

def count_weekend_days(file_path, shift_names):
    wb = load_workbook(file_path)
    # Search for sheetnames with 2024 in title
    sheets_to_search = [sheet for sheet in wb.sheetnames if '2024' in sheet]
    # Create dictionary {name: [number of WSat, number of WSun]}
    results_dict = {name: [0, 0] for name in shift_names}

    # Searching in sheets which have 2024 in name
    for sheet_name in sheets_to_search:
        sheet = wb[sheet_name]
        # Check column A if name matches shift_name in names_initials.csv
        for row in range(1, sheet.max_row + 1):
            person_name = sheet.cell(row=row, column=1).value
            if person_name in shift_names:
                # Look for working weekend days (wsat or wsun) from col B to the last to count them
                for col in range(2, sheet.max_column + 1):
                    shift_value = sheet.cell(row=row, column=col).value
                    if str(shift_value).lower() == 'wsat':
                        results_dict[person_name][0] += 1
                    elif str(shift_value).lower() == 'wsun':
                        results_dict[person_name][1] += 1
    # Close sheet
    wb.close()

    return results_dict