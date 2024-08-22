from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, NamedStyle, Alignment, Side, Border
from openpyxl.utils import get_column_letter
from collections import Counter
from tkinter import messagebox
import os
import config


def generate_shifts_sheet(days, people):
    min_day = datetime.strptime(days[0].date, '%d/%m/%Y')
    max_day = min_day
    for day in days:
        checked_date = datetime.strptime(day.date, '%d/%m/%Y')
        if checked_date < min_day:
            min_day = checked_date
        if checked_date > max_day:
            max_day = checked_date

    wb = Workbook()
    ws = wb.active

    # Write names
    alignment = Alignment(horizontal=config.WRITE_SHIFTS_CELLS_ALIGNMENT, vertical=config.WRITE_SHIFTS_CELLS_ALIGNMENT)
    date_style = NamedStyle(name='date_style', number_format=config.WRITE_SHIFTS_DATE_FORMAT, alignment=alignment)
    names_font = Font(name=config.WRITE_SHIFTS_FONT, size=config.WRITE_SHIFTS_FONT_SIZE_NAMES, bold=True)
    title_font = Font(name=config.WRITE_SHIFTS_FONT, size=config.WRITE_SHIFTS_FONT_SIZE_TITLE, bold=True)
    names_fill = PatternFill(start_color=config.WRITE_SHIFTS_NAMES_FILL,
                             end_color=config.WRITE_SHIFTS_NAMES_FILL,
                             fill_type=config.WRITE_SHIFTS_FILL_TYPE)
    holiday_fill = PatternFill(start_color=config.WRITE_SHIFTS_HOLIDAYS_FILL,
                               end_color=config.WRITE_SHIFTS_HOLIDAYS_FILL,
                               fill_type=config.WRITE_SHIFTS_FILL_TYPE)
    green_fill = PatternFill(start_color=config.WRITE_SHIFTS_WORKING_DAY_FILL,
                             end_color=config.WRITE_SHIFTS_WORKING_DAY_FILL,
                             fill_type=config.WRITE_SHIFTS_FILL_TYPE)
    blue_fill = PatternFill(start_color=config.WRITE_SHIFTS_WORKING_WEEKEND_FILL,
                            end_color=config.WRITE_SHIFTS_WORKING_WEEKEND_FILL,
                            fill_type=config.WRITE_SHIFTS_FILL_TYPE)

    for i, person in enumerate(people, start=config.NAMES_START_ROW):
        cell = ws.cell(row=i, column=config.NAMES_START_COL, value=person.shifts_name)
        cell.font = names_font
        cell.fill = names_fill

    # Adjust names-column width
    column_letter = get_column_letter(config.NAMES_START_COL)
    max_length = max(len(cell.value) if cell.value else 0 for cell in ws[get_column_letter(config.NAMES_START_COL)])
    adjusted_width = max_length + config.WRITE_SHIFTS_NAMES_WIDTH_MARGIN
    ws.column_dimensions[column_letter].width = adjusted_width

    def get_weekday_name(date_in):
        week_days = config.WRITE_SHIFTS_DISPLAYED_WEEKDAYS
        return week_days[date_in.weekday()]

    # Write dates
    write_date = min_day
    dates_col = config.NAMES_START_COL + 1

    while write_date <= max_day:
        col_letter = get_column_letter(dates_col)
        date_cell = ws.cell(row=config.WRITE_SHIFTS_DATES_ROW, column=dates_col, value=write_date)
        date_cell.style = date_style
        ws[f"{col_letter}5"] = get_weekday_name(write_date)
        ws[f"{col_letter}5"].alignment = alignment

        # Write non-working days
        for row in range(config.NAMES_START_ROW, config.NAMES_START_ROW + len(people)):
            cell = ws.cell(row=row, column=dates_col)
            ws.row_dimensions[row].height = config.WRITE_SHIFTS_NAMES_CELL_HEIGHT
            if write_date.weekday() < 5:  # Monday - Friday
                cell.value = config.WORKING_DAY_NAME
                cell.fill = green_fill
            elif write_date.weekday() == 5:  # Saturday
                cell.value = config.NON_WORKING_SATURDAY_NAME
            else:  # Sunday
                cell.value = config.NON_WORKING_SUNDAY_NAME
            cell.alignment = alignment

        write_date += timedelta(days=1)
        dates_col += 1

    # Write holidays

    for person in people:
        for date in person.holiday:
            formatted_date = datetime.strptime(date, '%d/%m/%Y')
            for col in range(2, ws.max_column + 1):
                if ws.cell(row=6, column=col).value == formatted_date:
                    for row in range(config.NAMES_START_ROW, config.NAMES_START_ROW + len(people)):
                        if ws.cell(row=row, column=1).value == person.shifts_name:
                            cell = ws.cell(row=row, column=col)
                            cell.value = config.HOLIDAY_NAME
                            cell.alignment = alignment
                            cell.fill = holiday_fill

    # Write assigned people
    for day in days:
        day_formatted_date = datetime.strptime(day.date, '%d/%m/%Y')
        for col in range(2, ws.max_column + 1):
            if ws.cell(row=6, column=col).value == day_formatted_date:
                for person_initials in day.assigned_people:
                    for person in people:
                        if person.initials == person_initials:
                            for row in range(config.NAMES_START_ROW, config.NAMES_START_ROW + len(people)):
                                if ws.cell(row=row, column=1).value == person.shifts_name:
                                    cell = ws.cell(row=row, column=col)
                                    if cell.value == config.NON_WORKING_SATURDAY_NAME:
                                        cell.value = config.WORKING_SATURDAY_NAME
                                    if cell.value == config.NON_WORKING_SUNDAY_NAME:
                                        cell.value = config.WORKING_SUNDAY_NAME
                                    if cell.value == config.HOLIDAY_NAME:
                                        cell.value = config.WORKING_HOLIDAY_NAME
                                    cell.alignment = alignment
                                    cell.fill = blue_fill

    # Write table borders
    table_thin_side = Side(style=config.WRITE_SHIFTS_TABLE_SIDE_STYLE)
    table_thin_border = Border(left=table_thin_side, right=table_thin_side, top=table_thin_side, bottom=table_thin_side)

    for row in ws.iter_rows(min_row=config.NAMES_START_ROW - 2,
                            max_row=config.NAMES_START_ROW + len(people),
                            min_col=config.NAMES_START_COL,
                            max_col=(max_day-min_day).days + 2
                            ):
        for cell in row:
            cell.border = table_thin_border

    # Check which month is most common in selected range to write it in title
    month_counter = Counter()
    current_date = min_day
    while current_date <= max_day:
        month_counter[current_date.month] += 1
        current_date += timedelta(days=1)
    most_common_month = month_counter.most_common(1)[0]
    current_year = datetime.now().year
    most_common_month_name = datetime(datetime.now().year, most_common_month[0], 1).strftime('%B')
    ws.cell(row=config.NAMES_START_ROW - 2, column=config.NAMES_START_COL, value="Team").font = title_font
    msg = f"{config.WRITE_SHIFTS_TITLE} - {most_common_month_name} {current_year}"
    title_cell = ws.cell(row=config.WRITE_SHIFTS_TITLE_ROW, column=config.WRITE_SHIFTS_TITLE_COL, value=msg)
    title_cell.font = title_font
    title_cell.alignment = alignment
    ws.merge_cells(config.WRITE_SHIFTS_TITLE_MERGE_CELLS)

    # Save
    file_path = config.WRITE_SHIFTS_FILE_NAME
    try:
        wb.save(file_path)
        # Open file
        os.startfile(file_path)
    except OSError:
        messagebox.showinfo("Error", "Cannot save! New shifts file is opened. Please close it and try again.")
