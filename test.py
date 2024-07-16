import pandas
from openpyxl import load_workbook
from datetime import datetime

class Person:
    def __init__(self, initials, holiday_name, shifts_name):
        self.initials = initials
        self.holiday_name = holiday_name
        self.shifts_name = shifts_name
        self.working_days = []
        self.unavailable = []


df = pandas.read_csv("names_initials.csv")
names_list = df.values.tolist()
people_in_team = []
for n in names_list:
    new_person = Person(n[0], n[1], n[2])
    people_in_team.append(new_person)

shift_file_names = [person.shifts_name for person in people_in_team]
print(shift_file_names)


def generate_unavailable_dict(file_path):
    # Wczytanie skoroszytu Excel
    wb = load_workbook(file_path)

    # Wybór arkusza "USP Scheduling"
    sheet = wb['USP scheduling']

    # Słownik na niedostępne osoby dla każdego dnia
    unavailable_dict = {}

    # Zakres kolumn i wierszy z danymi
    start_row = 5
    end_row = 736
    start_col = 1  # Kolumna A
    end_col = 60  # Kolumna daleka

    # Przeglądanie danych dla każdego dnia
    for row in range(start_row, end_row + 1):
        # Odczyt daty z kolumny A
        date_cell = sheet.cell(row=row, column=1)
        date_value = date_cell.value

        if date_value is None:
            continue

        # Konwersja daty do formatu "miesiąc/dzień/rok"
        date_str = date_value.strftime('%m/%d/%Y')  # Formatowanie daty z wiodącymi zerami

        # Sprawdzenie czy data należy do jednego z zadanego miesiąca, miesiąca poprzedzającego lub następnego
        month = date_value.month
        current_month = datetime.now().month
        if (month == current_month
                or month == current_month + 3
                or month == current_month + 2
                or month == current_month + 1
                or month == current_month - 1
                or month == current_month - 2
                or month == current_month - 3):
            # Dla danego dnia, sprawdzamy każdą kolumnę z imionami pracowników
            for col in range(start_col + 2, end_col + 1):  # Przesunięcie o 2, aby zacząć od kolumny C
                person_name = sheet.cell(row=4, column=col).value
                availability = sheet.cell(row=row, column=col).value

                if availability == 'Not available':
                    # Dodajemy osobę do listy niedostępnych dla danego dnia
                    if date_str not in unavailable_dict:
                        unavailable_dict[date_str] = []
                    unavailable_dict[date_str].append(person_name)

    # Zamknięcie skoroszytu
    wb.close()

    return unavailable_dict


# Przykład użycia funkcji
file_path = '02.2022-02.2023 Holiday calendar.xlsx'

unavailable_dict = generate_unavailable_dict(file_path)

# # Wyświetlenie wyników
# for day, people_list in unavailable_dict.items():
#     print(f"Day: {day}, Unavailable people: {', '.join(people_list)}")

print(unavailable_dict)


def update_unavailable_dates(unavailable_dict, people_in_team):
    for date, unavailable_people in unavailable_dict.items():
        for person in people_in_team:
            if person.holiday_name in unavailable_people:
                person.unavailable.append(date)


update_unavailable_dates(unavailable_dict, people_in_team)

for person in people_in_team:
    print(f"{person.holiday_name}: {person.unavailable}")